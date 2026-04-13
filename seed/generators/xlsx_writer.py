"""
XLSX writer — produces one workbook per month matching the real spreadsheet structure.
All values are pre-computed; no Excel formulas are written.
"""

import os
from openpyxl import Workbook
from seed.config import ROUTES


CONTROLE_HEADERS = [
    "Mês", "OS", "Tipo NF", "Data", "Empresa", "Solicitante", "Canal", "Agendador",
    "Placa", "Motorista",
    "Km Vazio", "Km Cheio", "Km Total", "Km Inicial", "Km Final",
    "H. Solicitada", "H. Saída", "H. Final", "T. Empresa", "T. Motorista",
    "Origem", "Destino", "Passageiro Principal", "Qtd Passageiros",
    "Tipo Serviço", "Tipo Viagem",
    "Pedágio", "Estacionamento", "Hora Parada", "Passagem",
    "Retorno", "Preço Corrida", "Adicional", "Total",
    "Km Noturno", "Km Feriado", "Km Domingo", "Preço/Km", "Custo/Passageiro",
    "Status",
]

CONTROLE_FIELD_MAP = [
    "month", "os_number", "nf_type", "ride_date", "company", "requester",
    "channel", "scheduler", "vehicle_plate", "driver",
    "km_empty", "km_loaded", "km_total", "km_initial", "km_final",
    "time_requested", "time_start", "time_end", "company_time", "driver_time",
    "origin", "destination", "main_passenger", "passenger_count",
    "service_type", "trip_type",
    "toll", "parking", "idle_time_billing", "transit_fare",
    "return_billing", "ride_price", "surcharges", "total",
    "night_km", "holiday_km", "sunday_km", "price_per_km", "cost_per_passenger",
    "status",
]


def _write_sheet(ws, headers: list, rows: list[dict], field_map: list = None):
    """Write headers in row 1, then one row per dict. field_map controls column order."""
    ws.append(headers)
    for row in rows:
        if field_map:
            ws.append([row.get(f) for f in field_map])
        else:
            ws.append(list(row.values()))


def write_workbook(
    month: str,
    rides: list[dict],
    empresa: list[dict],
    passageiro: list[dict],
    colaboradores: list[dict],
    frota: list[dict],
    folha: list[dict],
    financeiro: list[dict],
    frota_stats: list[dict],
    folga: list[dict],
    metas: list[dict],
    output_dir: str,
):
    """
    Write one XLSX file for the given month into output_dir.
    File name: VCTrans_YYYY-MM.xlsx
    """
    wb = Workbook()

    # Controle — active sheet by default
    ws_controle = wb.active
    ws_controle.title = "Controle"
    _write_sheet(ws_controle, CONTROLE_HEADERS, rides, CONTROLE_FIELD_MAP)

    # Passageiro
    ws_pass = wb.create_sheet("Passageiro")
    _write_sheet(ws_pass, list(passageiro[0].keys()), passageiro)

    # Empresa
    ws_emp = wb.create_sheet("Empresa")
    _write_sheet(ws_emp, list(empresa[0].keys()), empresa)

    # Colaboradores
    ws_col = wb.create_sheet("Colaboradores")
    _write_sheet(ws_col, list(colaboradores[0].keys()), colaboradores)

    # Frota
    ws_frota = wb.create_sheet("Frota")
    _write_sheet(ws_frota, list(frota[0].keys()), frota)

    # Região — route definitions from config
    ws_reg = wb.create_sheet("Região")
    ws_reg.append(["Origem", "Destino", "Km", "Tipo"])
    for route in ROUTES:
        ws_reg.append([route["origin"], route["destination"], route["km"], route["type"]])

    # Config — holiday dates referenced by billing rules
    ws_cfg = wb.create_sheet("Config")
    ws_cfg.append(["Chave", "Valor"])
    ws_cfg.append(["Feriado 1", "2026-01-01"])
    ws_cfg.append(["Feriado 2", "2026-02-16"])
    ws_cfg.append(["Feriado 3", "2026-02-17"])
    ws_cfg.append(["Feriado 4", "2026-03-06"])

    # Folha — driver payroll
    ws_folha = wb.create_sheet("Folha")
    _write_sheet(ws_folha, list(folha[0].keys()), folha)

    # Financeiro — company revenue summary
    ws_fin = wb.create_sheet("Financeiro")
    _write_sheet(ws_fin, list(financeiro[0].keys()), financeiro)

    # Frota Stats — vehicle performance
    ws_frota_stats = wb.create_sheet("Frota Stats")
    _write_sheet(ws_frota_stats, list(frota_stats[0].keys()), frota_stats)

    # Folga — driver schedule
    ws_folga = wb.create_sheet("Folga")
    _write_sheet(ws_folga, list(folga[0].keys()), folga)

    # Metas — KM goals
    ws_metas = wb.create_sheet("Metas")
    _write_sheet(ws_metas, list(metas[0].keys()), metas)

    # Frete — empty placeholder
    wb.create_sheet("Frete")

    filepath = os.path.join(output_dir, f"VCTrans_{month}.xlsx")
    wb.save(filepath)
    print(f"  Saved: {filepath}")
